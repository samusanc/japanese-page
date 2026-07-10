#!/usr/bin/env python3
"""
build_vocab.py — converts the Marugoto vocabulary/kanji xlsx files in
vocabulary/ into per-level JSON decks for the Royal Gamble card game.

  pip install openpyxl
  python scripts/build_vocab.py

Output: public/vocab/<level>.json with shape
  { "label": "...", "vocab": [{"ja","kana","es"}...], "kanji": [{"ja","kana"}...] }

- vocab entries: ja = kanji form (kana form if the book writes it in kana),
  kana = reading, es = Spanish translation
- kanji entries: ja = kanji word, kana = reading (matched against each other
  in the game's kanji mode; these lists carry no translations)
"""
import json
import re
import pathlib
import openpyxl

HERE = pathlib.Path(__file__).parent.parent.resolve()
SRC = HERE / "vocabulary"
OUT = HERE / "public" / "vocab"
OUT.mkdir(parents=True, exist_ok=True)

LEVELS = {
    "starter": {
        "label": "Starter (A1)",
        "vocab": ["starter_activities_vocabulary_index_ES.xlsx",
                  "starter_competences_vocabulary_index_ES.xlsx"],
        "kanji": ["starter_kanji_word_list.xlsx"],
    },
    "elementary1": {
        "label": "Elementary 1 (A2)",
        "vocab": ["elementary1_activities_vocabulary_index_ES.xlsx",
                  "elementary1_competences_vocabulary_index_ES.xlsx"],
        "kanji": ["elementary1_kanji_word_list.xlsx"],
    },
    "elementary2": {
        "label": "Elementary 2 (A2)",
        "vocab": ["elementary2_activities_vocabulary_index_ES.xlsx",
                  "elementary2_competences_vocabulary_index_ES.xlsx"],
        "kanji": ["elementary2_kanji_word_list.xlsx"],
    },
    "preintermediate": {
        "label": "Pre-Intermediate (A2/B1)",
        "vocab": ["preintermediate_vocabulary_index_ES.xlsx"],
        "kanji": ["pre-intermediate_kanji_word_list.xlsx"],
    },
    "intermediate1": {
        "label": "Intermediate 1 (B1)",
        "vocab": ["intermediate1_vocabulary_list_ES.xlsx"],
        "kanji": ["intermediate1_kanji_word_list.xlsx"],
    },
    "intermediate2": {
        "label": "Intermediate 2 (B1)",
        "vocab": ["intermediate2_vocabulary_list_ES.xlsx"],
        "kanji": ["intermediate2_kanji_word_list.xlsx"],
    },
}

WS = "　 \t"


def clean(v):
    if v is None:
        return ""
    return str(v).strip(WS).strip()


def parse_vocab_index(path):
    """Index format (starter/elementary/preintermediate): header row has
    IndexNo. / 語彙<かな> / 語彙<漢字> / … / スペイン語訳"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.worksheets[0]  # gojuon-ordered sheet
    rows = ws.iter_rows(values_only=True)
    header = None
    es_col = None
    out = []
    for r in rows:
        cells = [clean(c) for c in r]
        if header is None:
            if cells and cells[0] == "IndexNo.":
                header = cells
                es_col = next(i for i, c in enumerate(cells) if "スペイン語訳" in c)
            continue
        kana, kanji = cells[1], cells[2]
        es = cells[es_col] if es_col < len(cells) else ""
        if not kana or not es:
            continue
        # （…） in the kanji column = the book writes this word in kana
        ja = re.sub(r"[（）()]", "", kanji) or kana
        out.append({"ja": ja, "kana": kana, "es": es})
    return out


def parse_vocab_list(path):
    """Intermediate list format: TOPIC / PART / 見出し語 / アクセント / スペイン語訳.
    The アクセント column is the kana reading with accent digits embedded."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = None
    cols = {}
    out = []
    for r in rows:
        cells = [clean(c) for c in r]
        if header is None:
            if "見出し語" in cells:
                header = cells
                cols = {name: cells.index(name) for name in ("見出し語", "アクセント", "スペイン語訳")}
            continue
        ja = cells[cols["見出し語"]]
        kana = re.sub(r"[0-9?]", "", cells[cols["アクセント"]])
        es = cells[cols["スペイン語訳"]]
        if not ja or not es:
            continue
        out.append({"ja": ja, "kana": kana or ja, "es": es})
    return out


def parse_kanji_list(path):
    """Kanji word lists: header かんじ/漢字 + よみかた/読み方; topic separator
    rows have an empty kanji column."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    started = False
    out = []
    for r in rows:
        cells = [clean(c) for c in r]
        if not started:
            if len(cells) >= 4 and cells[2] in ("かんじ", "漢字"):
                started = True
            continue
        if len(cells) < 4:
            continue
        ja, kana = cells[2], cells[3]
        if not ja or not kana:
            continue
        out.append({"ja": ja, "kana": kana})
    return out


def dedupe(entries, key):
    seen = set()
    out = []
    for e in entries:
        k = key(e)
        if k in seen:
            continue
        seen.add(k)
        out.append(e)
    return out


def main():
    total_v = total_k = 0
    for level, spec in LEVELS.items():
        vocab = []
        for f in spec["vocab"]:
            p = SRC / f
            vocab += parse_vocab_list(p) if "vocabulary_list" in f else parse_vocab_index(p)
        vocab = dedupe(vocab, lambda e: (e["ja"], e["kana"], e["es"]))
        kanji = []
        for f in spec["kanji"]:
            kanji += parse_kanji_list(SRC / f)
        kanji = dedupe(kanji, lambda e: (e["ja"], e["kana"]))

        deck = {"label": spec["label"], "vocab": vocab, "kanji": kanji}
        out = OUT / f"{level}.json"
        out.write_text(json.dumps(deck, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        total_v += len(vocab)
        total_k += len(kanji)
        print(f"{level:16s} vocab={len(vocab):5d} kanji={len(kanji):4d} -> {out.name}")
    print(f"total: {total_v} vocab entries, {total_k} kanji words")


if __name__ == "__main__":
    main()
